import tkinter
from tkinter import colorchooser, messagebox, StringVar
import customtkinter
import settings #File containing settings
import globals  #File containing global variables
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from PIL import ImageGrab
from openpyxl.drawing.image import Image
import os
import helper_functions


class Material_Control_Panel:
    def __init__(self, window, row_placement:int, column_placement:int):
        # print("CLASS MATERIAL_CONTROL_PANEL INIT()")

        #Window where everything is placed
        self.program_window = window

        #Row/column placement in main program window
        self.row_placement = row_placement
        self.column_placement = column_placement

        self.material_control_panel_frame = self.create_material_control_panel()


    def create_material_control_panel(self):
        """Creates a control panel frame to control actions of materials"""
        # print("CREATE_MATERIAL_CONTROL_PANEL")

        #if material_control_frame has NOT been created before, create it
        if not hasattr(self, 'material_control_panel_frame'):
            #Create Frame from the control panel and place it within given window
            material_control_panel_frame = customtkinter.CTkFrame(
                master=self.program_window, 
                fg_color=settings.material_control_panel_background_color
            )
            material_control_panel_frame.grid(
                row=self.row_placement,
                column=self.column_placement,
                sticky="nsew",
                padx=(settings.material_control_panel_padding_left, settings.material_control_panel_padding_right),
                pady=(settings.material_control_panel_padding_top, settings.material_control_panel_padding_bottom)
            )

            #Define the row&column layout of the material_control_panel_frame
            material_control_panel_frame.columnconfigure(0, weight=50, uniform="group1")
            material_control_panel_frame.columnconfigure(1, weight=50, uniform="group1")

            material_control_panel_frame.rowconfigure(0, weight=50, uniform="group1")   
            material_control_panel_frame.rowconfigure(1, weight=50, uniform="group1")    

            

        #Create button to "add material" and place it
        add_material_button = customtkinter.CTkButton(
            master=material_control_panel_frame, 
            text="Add material", #"+" 
            fg_color=settings.material_control_panel_button_color,
            hover_color=settings.material_control_panel_button_hover_color,
            text_color=settings.material_control_panel_button_text_color,
            font=(settings.text_font, settings.material_control_panel_text_size),
            command=self.add_material
        )
        add_material_button.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(5,5),
            pady=(5,5)
        )

        #Create button to modify the materials
        modify_material_button = customtkinter.CTkButton(
            master=material_control_panel_frame, 
            text="Modify material",#"⚙️",
            font=(settings.text_font, settings.material_control_panel_text_size),
            fg_color=settings.material_control_panel_button_color,
            hover_color=settings.material_control_panel_button_hover_color,
            text_color=settings.material_control_panel_button_text_color, 
            command=self.modify_material
        )
        modify_material_button.grid(
            row=1,
            column=0,
            sticky="nsew",
            padx=(5,5),
            pady=(5,5)
        ) 

        #Reset values button
        reset_values_button = customtkinter.CTkButton(
            master=material_control_panel_frame,
            text="Reset materials",
            font=(settings.text_font, settings.material_control_panel_text_size),
            fg_color= settings.material_control_panel_button_color, 
            hover_color=settings.material_control_panel_button_hover_color, 
            text_color=settings.material_control_panel_button_text_color,
            command=self.reset_values
        )
        reset_values_button.grid(
            row=0, 
            column=1, 
            sticky="nsew", 
            padx=(5,5), 
            pady=(5,5)
        )           

        #Create button to export materials{} and stack to a excel file
        export_as_excel_button = customtkinter.CTkButton(
            master=material_control_panel_frame, 
            text="Export values to excel", 
            font=(settings.text_font, settings.material_control_panel_text_size),
            fg_color=settings.material_control_panel_button_color,
            hover_color=settings.material_control_panel_button_hover_color,
            text_color=settings.material_control_panel_button_text_color,
            command=self.export_to_excel
        )
        export_as_excel_button.grid(
            row=1,
            column=1,
            sticky="nsew",
            padx=(5,5),
            pady=(5,5)
        )

        return material_control_panel_frame
    

    def add_material(self):
        """Creates a popup window with value entries to add a new material in 'materials{}'"""
        # print("ADD_MATERIAL()")
        #Open up new program window
        self.add_material_window = tkinter.Toplevel(self.program_window)
        self.add_material_window.title("Add material")
        self.add_material_window.geometry(f"{settings.add_material_window_width}x{settings.add_material_window_height}")
        self.add_material_window.configure(bg=settings.add_material_window_background_color)

        #Define the row&column layout of the material_control_panel_frame
        self.add_material_window.columnconfigure(0, weight=34, uniform="group1")
        self.add_material_window.columnconfigure(1, weight=33, uniform="group1")
        self.add_material_window.columnconfigure(2, weight=33, uniform="group1")

        self.add_material_window.rowconfigure((0,1,2,3,4,5,6,7,8,9,10,11), weight=100, uniform="group1")


        #Create Labels and Entries for material properties 
        material_name_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Name", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        material_name_label.grid(
            row=0, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        #Name
        self.material_name_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.material_name_entry.grid(
            row=0, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #Thickness
        self.material_thickness_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Thickness [nm]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.material_thickness_label.grid(
            row=1, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.material_thickness_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.material_thickness_entry.grid(
            row=1, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #Indent
        self.material_indent_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Indent [nm]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.material_indent_label.grid(
            row=2, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.material_indent_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.material_indent_entry.grid(
            row=2, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #Color
        self.material_color_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Color", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.material_color_label.grid(
            row=3, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.material_color_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.material_color_entry.grid(
            row=3, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )


        #Modulus [GPa] value
        self.Modulus_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Modulus [GPa]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.Modulus_value_label.grid(
            row=4, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.Modulus_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.Modulus_value_entry.grid(
            row=4, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #CTE [ppm/deg] value
        self.CTE_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="CTE [ppm/deg]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.CTE_value_label.grid(
            row=5, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.CTE_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.CTE_value_entry.grid(
            row=5, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )


        #Density [kg/m3] value
        self.Density_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Density [kg/m3]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.Density_value_label.grid(
            row=6, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.Density_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.Density_value_entry.grid(
            row=6, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        
        #Stress_x [MPa] value
        self.Stress_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Stress_x [MPa]", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.Stress_value_label.grid(
            row=7, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.Stress_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.Stress_value_entry.grid(
            row=7, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )
            

        #Poisson value
        self.Poisson_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="Poisson", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.Poisson_value_label.grid(
            row=8, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.Poisson_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.Poisson_value_entry.grid(
            row=8, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #R0 value
        self.R0_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="R0", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.R0_value_label.grid(
            row=9, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.R0_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.R0_value_entry.grid(
            row=9, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )

        #R value
        self.R_value_label = customtkinter.CTkLabel(
            master=self.add_material_window, 
            text="R", 
            text_color=settings.add_material_window_text_color,
            fg_color=settings.add_material_window_background_color,
        )
        self.R_value_label.grid(
            row=10, 
            column=0, 
            sticky="e", 
            padx=(0,0),
            pady=(0,0)
        )

        self.R_value_entry = customtkinter.CTkEntry(
            master=self.add_material_window,
            fg_color = settings.add_material_window_entry_background_color,
            border_color=settings.add_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.add_material_window_entry_text_color,
            width=70,
            justify="center"
        )
        self.R_value_entry.grid(
            row=10, 
            column=1,
            sticky="ew",
            padx=(5,0),
            pady=(0,0)
        )


        #Find color buttom
        find_color_button = customtkinter.CTkButton(
            master=self.add_material_window,
            text="Find color",
            fg_color=settings.add_material_window_button_color,
            hover_color=settings.add_material_window_button_hover_color,
            text_color=settings.add_material_window_button_text_color,
            width=15,
            command=self.choose_color_for_add_material
        )
        find_color_button.grid(
            row=3,
            column=2,
            sticky="w",
            padx=(5,0),
            pady=(0,0)
        ) 

        #Confirm button
        confirm_button = customtkinter.CTkButton(
            master=self.add_material_window,
            text="Confirm",
            fg_color=settings.add_material_window_button_color,
            hover_color=settings.add_material_window_button_hover_color,
            text_color=settings.add_material_window_button_text_color,
            width=15,
            command=self.validate_add_material_inputs
        )
        confirm_button.grid(
            row=11,
            column=1,
            sticky="ew",
            padx=(0,0),
            pady=(3,0)
        )

        #Bind the "enter/return" button to call validate_add_material_inputs function when pressed
        self.add_material_window.bind('<Return>', lambda event: self.validate_add_material_inputs())


    def validate_add_material_inputs(self):
        """
        -Checks if the entry values from "add_material_window" are valid\n
        -Calls 'add_material_to_dictionary' if inputs are valid\n
        """
        #print("VALIDATE_INPUTS()")

        #Check if no entries are empty
        if(not self.material_name_entry.get()):
            messagebox.showerror("ERROR", "'Name' can not be empty", parent=self.add_material_window)
            self.add_material_window.lift()
            return
        
        #Check if material_name already exists in materials dictionary
        if(self.material_name_entry.get() in globals.materials):
            messagebox.showerror("ERROR", "'Name' already exists", parent=self.add_material_window)
            return


        #If "Thickness [nm]" value is entered, check if it is valid
        if(self.material_thickness_entry.get() != ""):
            thickness_value = globals.app.convert_decimal_string_to_float(self.material_thickness_entry.get())
            #If thickness is negative
            if(thickness_value < 0):
                messagebox.showerror("ERROR", "'Thickness' value can not be negative", parent=self.add_material_window)
                return
            #If thickness is not integer or float
            if(isinstance(thickness_value, bool) and thickness_value == False):
                messagebox.showerror("ERROR", "'Thickness' value has to be an integer or decimal number", parent=self.add_material_window)
                return

        #if "indent" value is entered, check if it is integer
        if(self.material_indent_entry.get() != ""):
            indent_value = globals.app.convert_decimal_string_to_float(self.material_indent_entry.get()) 
            if(isinstance(indent_value, bool) and indent_value == False):            
                messagebox.showerror("ERROR", "'Indent' value has to be an integer or decimal number", parent=self.add_material_window)
                return

        #Color
        if(not self.material_color_entry.get()):
            messagebox.showerror("ERROR", "'Color' can not be empty", parent=self.add_material_window)
            self.add_material_window.lift()
            return
        
        #Check if input color is valid
        if(self.is_valid_color(self.material_color_entry.get()) == False):
            messagebox.showerror("ERROR", "Given color is not valid.\nValue must be valid string or hex-value ('#123456)", parent=self.add_material_window)
            self.add_material_window.lift()
            return

        #If "Modulus" value is entered, check if it is integer
        if(self.Modulus_value_entry.get() != ""):
            modulus_value = globals.app.convert_decimal_string_to_float(self.Modulus_value_entry.get()) 
            if(isinstance(modulus_value, bool) and modulus_value == False):            
                messagebox.showerror("ERROR", "'Modulus [GPa]' value has to be an integer or decimal", parent=self.add_material_window)
                return
            if(modulus_value < 0):
                messagebox.showerror("ERROR", "'Modulus [GPa]' value can not be a negative number", parent=self.add_material_window)
                return

        #If "CTE" value is entered, check if it is integer
        if(self.CTE_value_entry.get() != ""):
            CTE_value = globals.app.convert_decimal_string_to_float(self.CTE_value_entry.get())
            if(isinstance(CTE_value, bool) and CTE_value == False):
                messagebox.showerror("ERROR", "'CTE [ppm/deg]' value has to be an integer or decimal", parent=self.add_material_window)
                return
            if(CTE_value < 0):
                messagebox.showerror("ERROR", "'CTE [ppm/deg]' value can not be a negative number", parent=self.add_material_window)
                return
        
        #If "Density" value is entered, check if it is integer
        if(self.Density_value_entry.get() != ""):
            density_value = globals.app.convert_decimal_string_to_float(self.Density_value_entry.get()) 
            if(isinstance(density_value, bool) and density_value == False):            
                messagebox.showerror("ERROR", "'Density [kg/m3]' value has to be an integer or decimal", parent=self.add_material_window)
                return
            if(density_value < 0):
                messagebox.showerror("ERROR", "'Density [kg/m3]' value can not be a negative number", parent=self.add_material_window)
                return

        #If 'Stress_x [MPa]' value is entered, check if it is integer 
        if(self.Stress_value_entry.get() != ""):
            stress_value = globals.app.convert_decimal_string_to_float(self.Stress_value_entry.get()) 
            if(isinstance(stress_value, bool) and stress_value == False):            
                messagebox.showerror("ERROR", "'Stress_x [MPa]' value has to be an integer or decimal", parent=self.add_material_window)
                return

        #If 'Poisson' value is entered, check if it is integer
        if(self.Poisson_value_entry.get() != ""):
            poisson_value = globals.app.convert_decimal_string_to_float(self.Poisson_value_entry.get())
            if(isinstance(poisson_value, bool) and poisson_value == False):            
                messagebox.showerror("ERROR", "'Poisson' value has to be an integer or decimal", parent=self.add_material_window)
                return
            if(poisson_value < 0):
                messagebox.showerror("ERROR", "'Poisson' value can not be a negative number", parent=self.add_material_window)
                return

        #If 'R0' value is entered, check if it is integer
        if(self.R0_value_entry.get() != ""):
            R0_value = globals.app.convert_decimal_string_to_float(self.R0_value_entry.get()) 
            if(isinstance(R0_value, bool) and R0_value == False):            
                messagebox.showerror("ERROR", "'R0' value has to be an integer or decimal", parent=self.add_material_window)
                return

        #If 'R' value is entered, check if it is integer
        if(self.R_value_entry.get() != ""):
            R_value = globals.app.convert_decimal_string_to_float(self.R_value_entry.get()) 
            if(isinstance(R_value, bool) and R_value == False):            
                messagebox.showerror("ERROR", "'R' value has to be an integer or decimal", parent=self.add_material_window)
                return


        #All inputs have been validated, add it to dictionary
        self.add_material_to_dictionary()

        #Destroy window
        self.add_material_window.destroy()
        
        #Update material_adjustment_panel
        globals.material_adjustment_panel.create_material_adjustment_panel()

        #Re-draw the material stack
        globals.layer_stack_canvas.draw_material_stack()


    def choose_color_for_add_material(self):
        """Opens up a color palette window and inserts the color code to 'material_color_entry' in 'add_material_window'"""
        # print("CHOOSE_COLOR_FOR_ADD_MATERIAL()")
        color_code = colorchooser.askcolor(title="Choose a color")[1]
        if(color_code):
            #Delete existing string in entry and insert new one
            self.material_color_entry.delete(0)
            self.material_color_entry.insert(0, color_code)
        #Lift the add_material_window to the front
        self.add_material_window.lift()

    
    def is_valid_color(self, color):
        """Returns True if given color string is a valid color. Return False if invalid"""
        # print("IS_VALID_COLOR()")
        #Check if color is accepted by tkinter
        try:
            self.program_window.winfo_rgb(color)
            return True
        
        except:
            #Check if color is valid hexadecimal value
            if(type(color)==str and color.startswith('#') and len(color)==7):
                return True
            else:
                return False
    

    def add_material_to_dictionary(self):
        """Adds values from entries to 'materials' dictionary"""
        # print("ADD_MATERIAL_TO_DICTIONARY")

        #If some entries in "add_material_window" are not set, the values is automaticly set to default value
        #NAME
        if(not self.material_name_entry.get()):
            material_name = tkinter.StringVar(value="No name")
        else:
            material_name = tkinter.StringVar(value=self.material_name_entry.get())

        #LAYER
        layer_value = tkinter.IntVar(value=len(globals.materials)+1)

        #THICKNESS
        if(not self.material_thickness_entry.get()):
            material_thickness = tkinter.DoubleVar(value=0)
        else:
            material_thickness = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.material_thickness_entry.get()))

        #UNIT
        unit_value = tkinter.StringVar(value="nm")

        #INDENT
        if(not self.material_indent_entry.get()):
            material_indent = tkinter.DoubleVar(value=0)
        else:
            material_indent = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.material_indent_entry.get()))
        
        #COLOR
        if(not self.material_color_entry.get()):
            material_color = tkinter.StringVar(value="white")
        else:
            material_color = tkinter.StringVar(value=self.material_color_entry.get())

        #STATUS
        status_value = tkinter.StringVar(value="active")

        #Modulus [GPa] VALUE
        if( not self.Modulus_value_entry.get()):
            Modulus_value = tkinter.DoubleVar(value=0)
        else:
            Modulus_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.Modulus_value_entry.get()))

        #CTE [ppm/deg] VALUE
        if(not self.CTE_value_entry.get()):
            CTE_value = tkinter.DoubleVar(value=0)
        else:
            CTE_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.CTE_value_entry.get()))

        #Density [kg/m3] VALUE
        if(self.Density_value_entry.get() == ""):
            Density_value = tkinter.DoubleVar(value=0)
        else:
            Density_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.Density_value_entry.get()))
        
        #Stress_x [MPa] VALUE
        if(self.Stress_value_entry.get() == ""):
            Stress_value = tkinter.DoubleVar(value=0)
        else:
            Stress_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.Stress_value_entry.get()))
        
        #POISSON VALUE
        if(self.Poisson_value_entry.get() == ""):
            Poisson_value = tkinter.DoubleVar(value=0)
        else:
            Poisson_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.Poisson_value_entry.get()))

        #R0 VALUE
        if(self.R0_value_entry.get() == ""):
            R0_value = tkinter.DoubleVar(value=0)
        else:
            R0_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.R0_value_entry.get()))

        #R VALUE
        if(self.R_value_entry.get() == ""):
            R_value = tkinter.DoubleVar(value=0)
        else:
            R_value = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.R_value_entry.get()))


        #ADD TRACES TO ALL VARIABLES
        material_name.trace_add("write", lambda *args, identifier="Name": globals.app.variable_updated(identifier))
        layer_value.trace_add("write", lambda *args, identifier="Layer": globals.app.variable_updated(identifier))
        material_thickness.trace_add("write", lambda *args, identifier="Thickness [nm]": globals.app.variable_updated(identifier))
        unit_value.trace_add("write", lambda *args, identifier="Unit": globals.app.variable_updated(identifier))
        material_indent.trace_add("write", lambda *args, identifier="Indent [nm]": globals.app.variable_updated(identifier))
        material_color.trace_add("write", lambda *args, identifier="Color": globals.app.variable_updated(identifier))
        status_value.trace_add("write", lambda *args, identifier="Status": globals.app.variable_updated(identifier))
        Modulus_value.trace_add("write", lambda *args, identifier="Modulus [GPa]": globals.app.variable_updated(identifier))
        CTE_value.trace_add("write", lambda *args, identifier="CTE [ppm/deg]": globals.app.variable_updated(identifier))
        Density_value.trace_add("write", lambda *args, identifier="Density [kg/m3]": globals.app.variable_updated(identifier))
        Stress_value.trace_add("write", lambda *args, identifier="Stress_x [MPa]": globals.app.variable_updated(identifier))
        Poisson_value.trace_add("write", lambda *args, identifier="Poisson": globals.app.variable_updated(identifier))
        R0_value.trace_add("write", lambda *args, identifier="R0": globals.app.variable_updated(identifier))
        R_value.trace_add("write", lambda *args, identifier="R": globals.app.variable_updated(identifier))
        

        #Add values to dictionary
        info = {
            "Name": material_name,
            "Layer": layer_value,
            "Thickness [nm]": material_thickness,
            "Unit": unit_value,
            "Indent [nm]": material_indent,
            "Color": material_color,
            "Status": status_value,
            "Modulus [GPa]": Modulus_value,
            "CTE [ppm/deg]": CTE_value,
            "Density [kg/m3]": Density_value,
            "Stress_x [MPa]": Stress_value,
            "Poisson": Poisson_value,
            "R0": R0_value,
            "R": R_value,
            "Label_name_id": None,
            "Delete_material_button_id": None,
            "Move_down_button_id": None,
            "Move_up_button_id": None,
            "Rectangle_id": None,
            "Text_id": None,
            "Text_bbox_id" : None,
            "Line_id": None,
            "Entry_id": None,
            "Slider_id": None,
            "Checkbox_id": None,
            "Indent_text_id": None,
            "Indent_text_bbox_id": None,
            "Indent_line_id": None,
            "Indent_arrow_pointer_id": None
        }
        
        #Put "info" dictionary into self.materials dictionary
        globals.materials[str(self.material_name_entry.get())] = info

        #Sort the materials dictionary
        globals.app.sort_dictionary()


    def modify_material(self):
        """Creates a new 'modify_material' window where the user can change the attributes of each material"""
        # print("MODIFY_MATERIAL()")

        #Create dictionary to contain ALL entries
        self.entry_dictionary = {}

        #Open up new program window
        self.modify_material_window = tkinter.Toplevel(self.program_window)
        self.modify_material_window.title("Modify materials")
        self.modify_material_window.geometry(f"{settings.modify_material_window_width}x{settings.modify_material_window_height}")
        self.modify_material_window.configure(bg=settings.modify_material_window_background_color)

        #Define the row&column layout of the self.modify_material_window
        self.modify_material_window.columnconfigure(0, weight=100, uniform="group1")
        self.modify_material_window.rowconfigure(0, weight=100, uniform="group1")

        #Bind the "renter/return" button to call "confirm_material_changes" function when pressed
        self.modify_material_window.bind('<Return>', lambda event: self.validate_modify_material_inputs())

        #Create a scrollable frame for entries to each material
        scrollable_frame = customtkinter.CTkScrollableFrame(
            master=self.modify_material_window,
            fg_color = settings.modify_material_window_scrollable_frame_background_color,
        )
        scrollable_frame.grid(
            row=0,
            column=0,
            sticky="nwse",
            padx=(0,0),
            pady=(0,0)
        )

        #Define the row&column layout of the scrollable_frame
        scrollable_frame.columnconfigure(0, weight=10, uniform="group1")    #name
        scrollable_frame.columnconfigure(1, weight=8, uniform="group1")     #Thickness
        scrollable_frame.columnconfigure(2, weight=5, uniform="group1")     #unit
        scrollable_frame.columnconfigure(3, weight=8, uniform="group1")     #Indent
        scrollable_frame.columnconfigure(4, weight=9, uniform="group1")    #Color
        scrollable_frame.columnconfigure(5, weight=10, uniform="group1")    #Modulus
        scrollable_frame.columnconfigure(6, weight=10, uniform="group1")    #CTE
        scrollable_frame.columnconfigure(7, weight=10, uniform="group1")    #Density
        scrollable_frame.columnconfigure(8, weight=10, uniform="group1")    #Stress
        scrollable_frame.columnconfigure(9, weight=8, uniform="group1")    #Poisson
        scrollable_frame.columnconfigure(10, weight=5, uniform="group1")    #R0
        scrollable_frame.columnconfigure(11, weight=5, uniform="group1")    #R

        scrollable_frame.rowconfigure(0, weight=10, uniform="group1")

        #Create headlines in the scrollable frame
        #Name
        name_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Name",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color=settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color,
        )
        name_label.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )

        #Thickness  
        thickness_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Thickness [nm]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        thickness_label.grid(
            row=0,
            column=1,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Unit
        unit_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Unit",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        unit_label.grid(
            row=0,
            column=2,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Indent[nm]
        indent_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Indent[nm]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        indent_label.grid(
            row=0,
            column=3,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Color
        color_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Color",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        color_label.grid(
            row=0,
            column=4,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Modulus [GPa]
        modulus_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Modulus [GPa]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        modulus_label.grid(
            row=0,
            column=5,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #CTE [ppm/deg]
        CTE_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="CTE [ppm/deg]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        CTE_label.grid(
            row=0,
            column=6,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Density [kg/m3]
        density_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Density [kg/m3]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        density_label.grid(
            row=0,
            column=7,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
        #Stress_x [MPa]
        stress_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Stress_x [MPa]",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        stress_label.grid(
            row=0,
            column=8,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )

        #Poisson
        poisson_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="Poisson",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        poisson_label.grid(
            row=0,
            column=9,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )

        #R0
        R0_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="R0",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        R0_label.grid(
            row=0,
            column=10,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )

        #R
        R_label = customtkinter.CTkLabel(
            master=scrollable_frame,
            text="R",
            font=(settings.modify_material_window_text_font, settings.modify_material_window_text_size, "bold"),
            text_color = settings.modify_material_window_text_color,
            fg_color=settings.modify_material_window_scrollable_frame_background_color
        )
        R_label.grid(
            row=0,
            column=11,
            sticky="nsew",
            padx=(0,0),
            pady=(0,0)
        )
        
                
        #Create entries for all materials
        row_counter = 1
        for material in dict(reversed(globals.materials.items())):
            inner_dictionary = {}
            
            #Create entries for each category in materials

            #MATERIAL NAME
            material_name_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Name"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            material_name_entry.grid(
                row=row_counter,
                column=0,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            #Add entry to dictionary
            inner_dictionary["Name_entry"] = material_name_entry

            #THICKNESS
            thickness_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Thickness [nm]"].get()),
                fg_color=settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            thickness_entry.grid(
                row=row_counter,
                column=1,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Thickness_entry"] = thickness_entry

            
            #UNIT
            unit_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Unit"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            unit_entry.grid(
                row=row_counter,
                column=2,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Unit_entry"] = unit_entry

            
            #INDENT [NM]
            indent_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Indent [nm]"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            indent_entry.grid(
                row=row_counter,
                column=3,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Indent_entry"] = indent_entry

            
            #COLOR
            color_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Color"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            color_entry.grid(
                row=row_counter,
                column=4,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Color_entry"] = color_entry

            
            #MODULUS [GPa]
            modulus_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Modulus [GPa]"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            modulus_entry.grid(
                row=row_counter,
                column=5,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Modulus_entry"] = modulus_entry

            
            #CTE [ppm/deg]
            CTE_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["CTE [ppm/deg]"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            CTE_entry.grid(
                row=row_counter,
                column=6,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["CTE_entry"] = CTE_entry

            
            #DENSITY [kg/m3]
            density_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Density [kg/m3]"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            density_entry.grid(
                row=row_counter,
                column=7,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Density_entry"] = density_entry

            
            #STRESS_X [MPa]
            stress_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Stress_x [MPa]"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            stress_entry.grid(
                row=row_counter,
                column=8,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Stress_entry"] = stress_entry

            #POISSON
            poisson_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["Poisson"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            poisson_entry.grid(
                row=row_counter,
                column=9,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["Poisson_entry"] = poisson_entry


            #R0
            R0_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=tkinter.StringVar(value=globals.materials[material]["R0"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            R0_entry.grid(
                row=row_counter,
                column=10,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["R0_entry"] = R0_entry


            #R
            R_entry = customtkinter.CTkEntry(
                master=scrollable_frame,
                textvariable=StringVar(value=globals.materials[material]["R"].get()),
                fg_color = settings.modify_material_window_entry_background_color,
                border_color=settings.modify_material_window_entry_border_color,
                border_width=0.4,
                text_color=globals.materials[material]["Color"].get(),
                justify="center"
            )
            R_entry.grid(
                row=row_counter,
                column=11,
                sticky="",
                padx=(0,0),
                pady=(0,0)
            )
            inner_dictionary["R_entry"] = R_entry


            self.entry_dictionary[material] = inner_dictionary

            row_counter += 1
        
        #Find color buttom
        find_color_button = customtkinter.CTkButton(
            master=self.modify_material_window,
            text="Find color",
            fg_color=settings.modify_material_window_button_color,
            hover_color=settings.modify_material_window_button_hover_color,
            text_color=settings.modify_material_window_button_text_color,
            command=self.choose_color_for_modify_material
        )
        find_color_button.grid(
            row=1,
            column=0,
            sticky="w",
            padx=(5,0),
            pady=(0,0)
        )
        
        
        #SHOW SELECTED COLOR ENTRY
        self.color_finder_entry = customtkinter.CTkEntry(
            master=self.modify_material_window,
            placeholder_text="No color selected",
            fg_color = settings.modify_material_window_entry_background_color,
            border_color=settings.modify_material_window_entry_border_color,
            border_width=0.4,
            text_color=settings.modify_material_window_entry_text_color,
            justify="center"
        )
        self.color_finder_entry.grid(
            row=2,
            column=0,
            sticky="w",
            padx=(5,0),
            pady=(0,0)
        )

        #CONFIRM CHANGES BUTTON
        confirm_button = customtkinter.CTkButton(
            master=self.modify_material_window,
            text="Confirm changes",
            fg_color=settings.modify_material_window_button_color,
            hover_color=settings.modify_material_window_button_hover_color,
            text_color=settings.modify_material_window_button_text_color,
            command=self.validate_modify_material_inputs
        )
        confirm_button.grid(
            row=3,
            column=0,
            sticky="",
            padx=(0,0),
            pady=(0,5)
        )
    

    def choose_color_for_modify_material(self):
        """Opens a color palette window and inserts the color code to 'self.color_finder_entry' in modify_materials_window'"""
        # print("CHOOSE_COLOR_FOR_MODIFY_MATERIAL()")
        color_code = colorchooser.askcolor(title="Choose a color")[1]

        if(color_code):
            self.color_finder_entry.delete(0)
            self.color_finder_entry.insert(0, color_code)
        
        #Pull modify_material_window to the front
        self.modify_material_window.lift()


    def validate_modify_material_inputs(self):
        """
        Goes through all entries in self.entry_dictionary and validates them
        """
        # print("VALIDATE_MODIFY_MATERIAL_INPUTS()")
        
        for material in self.entry_dictionary:
            #NAME
            if(self.entry_dictionary[material]["Name_entry"].get() == ""):
                messagebox.showerror("ERROR", f"'Material' for '{material}' can not be empty", parent=self.modify_material_window)
                self.modify_material_window.lift()
                return


            #THICKNESS
            if(self.entry_dictionary[material]["Thickness_entry"].get() != ""):
                thickness_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Thickness_entry"].get()) 
                if(isinstance(thickness_value, bool) and thickness_value == False):            
                    messagebox.showerror("ERROR", f"'Thickness' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(thickness_value < 0):
                    messagebox.showerror("ERROR", f"'Thickness' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If thickness box is empty
            else:
                messagebox.showerror("ERROR", f"'Thickness' value for '{material}' can not be empty", parent=self.modify_material_window)
                return
        
            #UNIT
            #If unit box is empty
            if(self.entry_dictionary[material]["Unit_entry"].get() == ""):
                messagebox.showerror("ERROR", f"'Unit' value for '{material}' can not be empty", parent=self.modify_material_window)
                self.modify_material_window.lift()
                return


            #INDENT [NM]
            if(self.entry_dictionary[material]["Indent_entry"].get() != ""):
                indent_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Indent_entry"].get()) 
                if(isinstance(indent_value, bool) and indent_value == False):            
                    messagebox.showerror("ERROR", f"'Indent[nm]' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(indent_value < 0):
                    messagebox.showerror("ERROR", f"'Indent' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If Indent box is empty
            else:
                messagebox.showerror("ERROR", f"'Indent' value for '{material}' can not be empty", parent=self.modify_material_window)
                return
            

            #COLOR
            #If entry is empty
            if(self.entry_dictionary[material]["Color_entry"].get() == ""):
                messagebox.showerror("ERROR", f"'Color' for '{material}' can not be empty", parent=self.modify_material_window)
                self.modify_material_window.lift()
                return
            #If entry is not empty
            else:
                #Check if input color is valid
                if(self.is_valid_color(self.entry_dictionary[material]["Color_entry"].get()) == False):
                    messagebox.showerror("ERROR", f"Given 'Color' for '{material}' is not valid.\nValue must be valid string or hex-value ('#123456)", parent=self.modify_material_window)
                    self.modify_material_window.lift()
                    return


            #MODULUS [GPa]
            if(self.entry_dictionary[material]["Modulus_entry"].get() != ""):
                modulus_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Modulus_entry"].get()) 
                if(isinstance(modulus_value, bool) and modulus_value == False):            
                    messagebox.showerror("ERROR", f"'Modulus[GPa]' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(modulus_value < 0):
                    messagebox.showerror("ERROR", f"'Modulus[GPa]' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If modulus entry is empty
            else:
                messagebox.showerror("ERROR", f"'Modulus[GPa]' value for '{material}' can not be empty", parent=self.modify_material_window)
                return

            

            #CTE [ppm/deg]
            if(self.entry_dictionary[material]["CTE_entry"].get() != ""):
                cte_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["CTE_entry"].get()) 
                if(isinstance(cte_value, bool) and cte_value == False):            
                    messagebox.showerror("ERROR", f"'CTE [ppm/deg]' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(cte_value < 0):
                    messagebox.showerror("ERROR", f"'CTE [ppm/deg]' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If CTE entry is empty
            else:
                messagebox.showerror("ERROR", f"'CTE [ppm/deg]' value for '{material}' can not be empty", parent=self.modify_material_window)
                return

            
            #DENSITY [kg/m3]
            if(self.entry_dictionary[material]["Density_entry"].get() != ""):
                density_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Density_entry"].get()) 
                if(isinstance(density_value, bool) and density_value == False):            
                    messagebox.showerror("ERROR", f"'Density [kg/m3]' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(density_value < 0):
                    messagebox.showerror("ERROR", f"'Density [kg/m3]' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If CTE entry is empty
            else:
                messagebox.showerror("ERROR", f"'Density [kg/m3]' value for '{material}' can not be empty", parent=self.modify_material_window)
                return


            #STRESS_X [MPa]
            if(self.entry_dictionary[material]["Stress_entry"].get() != ""):
                stress_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Stress_entry"].get()) 
                if(isinstance(stress_value, bool) and stress_value == False):            
                    messagebox.showerror("ERROR", f"'Stress [MPa]' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
            #If STRESS_X entry is empty
            else:
                messagebox.showerror("ERROR", f"'Stress [MPa]' value for '{material}' can not be empty", parent=self.modify_material_window)
                return

        
            #POISSON
            if(self.entry_dictionary[material]["Poisson_entry"].get() != ""):
                poisson_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Poisson_entry"].get()) 
                if(isinstance(poisson_value, bool) and poisson_value == False):            
                    messagebox.showerror("ERROR", f"'Poisson' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
                if(poisson_value < 0):
                    messagebox.showerror("ERROR", f"'Poisson' value for '{material}' can not be a negative number", parent=self.modify_material_window)
                    return
            #If poisson entry is empty
            else:
                messagebox.showerror("ERROR", f"'Poisson' value for '{material}' can not be empty", parent=self.modify_material_window)
                return


            #R0
            if(self.entry_dictionary[material]["R0_entry"].get() != ""):
                R0_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R0_entry"].get()) 
                if(isinstance(R0_value, bool) and R0_value == False):            
                    messagebox.showerror("ERROR", f"'R0' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
            #If R0 entry is empty
            else:
                messagebox.showerror("ERROR", f"'R0' value for '{material}' can not be empty", parent=self.modify_material_window)
                return



            #R
            if(self.entry_dictionary[material]["R_entry"].get() != ""):
                R_value = helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R_entry"].get()) 
                if(isinstance(R_value, bool) and R_value == False):            
                    messagebox.showerror("ERROR", f"'R' value for '{material}' has to be an integer or decimal", parent=self.modify_material_window)
                    return
            #If R entry is empty
            else:
                messagebox.showerror("ERROR", f"'R' value for '{material}' can not be empty", parent=self.modify_material_window)
                return


        self.confirm_material_changes()

    
    def confirm_material_changes(self):
        """
        -Makes changes to the materials{} dictionary based on the values given by user\n
        -If the user modified the 'material_name' then the main key in materials{} is changed to the new name 
        """
        # print("CONFIRM_MATERIAL_CHANGES()")

        for material in self.entry_dictionary:
            #Change the main key in materials{} dictionary
            new_key = self.entry_dictionary[material]["Name_entry"].get()
            globals.materials[new_key] = globals.materials.pop(material)   

            #Replace values in dictionary 
            # globals.materials[new_key]["Name"] = tkinter.StringVar(value=self.entry_dictionary[material]["Name_entry"].get())
            # globals.materials[new_key]["Thickness [nm]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Thickness_entry"].get()))  
            # globals.materials[new_key]["Unit"] = tkinter.StringVar(value=self.entry_dictionary[material]["Unit_entry"].get())  
            # globals.materials[new_key]["Indent [nm]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Indent_entry"].get())) 
            # globals.materials[new_key]["Color"] = tkinter.StringVar(value=self.entry_dictionary[material]["Color_entry"].get())  
            # globals.materials[new_key]["Modulus [GPa]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Modulus_entry"].get()))  
            # globals.materials[new_key]["CTE [ppm/deg]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["CTE_entry"].get()))
            # globals.materials[new_key]["Density [kg/m3]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Density_entry"].get()))
            # globals.materials[new_key]["Stress_x [MPa]"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Stress_entry"].get()))
            # globals.materials[new_key]["Poisson"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Poisson_entry"].get()))
            # globals.materials[new_key]["R0"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R0_entry"].get()))
            # globals.materials[new_key]["R"] = tkinter.DoubleVar(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R_entry"].get()))


            globals.materials[new_key]["Name"].set(self.entry_dictionary[material]["Name_entry"].get())
            globals.materials[new_key]["Thickness [nm]"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Thickness_entry"].get()))  
            globals.materials[new_key]["Unit"].set(self.entry_dictionary[material]["Unit_entry"].get())  
            globals.materials[new_key]["Indent [nm]"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Indent_entry"].get())) 
            globals.materials[new_key]["Color"].set(self.entry_dictionary[material]["Color_entry"].get())  
            globals.materials[new_key]["Modulus [GPa]"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Modulus_entry"].get()))  
            globals.materials[new_key]["CTE [ppm/deg]"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["CTE_entry"].get()))
            globals.materials[new_key]["Density [kg/m3]"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Density_entry"].get()))
            globals.materials[new_key]["Stress_x [MPa]"].set(value=helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Stress_entry"].get()))
            globals.materials[new_key]["Poisson"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["Poisson_entry"].get()))
            globals.materials[new_key]["R0"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R0_entry"].get()))
            globals.materials[new_key]["R"].set(helper_functions.convert_decimal_string_to_float(self.entry_dictionary[material]["R_entry"].get()))


        #Sort materials dictionary
        globals.app.sort_dictionary()

        #Destroy modify_material_window
        self.modify_material_window.destroy()

        #Re-create the material_adjustment_panel
        globals.material_adjustment_panel.create_material_adjustment_panel()

        #Redraw the material stack
        globals.layer_stack_canvas.draw_material_stack()



    def reset_values(self):
        """Repopulates globals.materials dictionary with values from the excel file and recreates the material_adjustment_panel """
       
        # print("RESET_VALUES")

        excel_file = "Materials.xlsx"

        #If there is a "materials" file in the folder, read it and reset the thickness values of each material
        if(os.path.isfile(excel_file)):
            #Delete all widgets in material_adjustment_panel_frame
            for widget in globals.material_adjustment_panel.material_adjustment_panel_frame.winfo_children():
                widget.destroy()
            
            del globals.material_adjustment_panel.material_headline
            del globals.material_adjustment_panel.slider_label

            
            #Clear the existing globals.materials
            globals.materials.clear()

            #Reload the values from the excel file in to the dictionary
            globals.app.load_materials_from_excel()

            #Redraw the material stack
            globals.layer_stack_canvas.draw_material_stack()

            #If graphs has been created, redraw them
            if(globals.graph_canvas != None):
                globals.graph_canvas.draw_z_tip_is_graph()
                globals.graph_canvas.draw_stoney_graph()

            #Recreate the material_adjustment_panel
            globals.material_adjustment_panel.create_material_adjustment_panel()

        else:
            messagebox.showerror("Error", "Can not reset values because there is no 'materials.xlsx' file to fetch original values from")


    def export_to_excel(self):
        """Saves the values from materials{} to an excel file and places a screenshot of the current stack in the excel file"""
        # print("EXPORT_TO_EXCEL()")

        #Create an filename and a workbook to contain data
        filename = "exported_materials.xlsx"

        #Create main folder if it does not already exist
        main_folder = "exports"
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder if it does not already exist
        sub_folder = "excel"
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        

        #Create path for file to be saved in
        file_path = os.path.join(f"{main_folder}/{sub_folder}/{filename}")

        workbook = Workbook()

        # Optionally, rename the default sheet
        sheet = workbook.active
        # sheet.title = ""

        #Create header cells
        sheet["A1"] = "Material"          # Add a new header in column A row 1
        sheet["B1"] = "Thickness [nm]"
        sheet["C1"] = "Unit"
        sheet["D1"] = "Indent [nm]"
        sheet["E1"] = "Color"
        sheet["F1"] = "Modulus [GPa]"
        sheet["G1"] = "CTE [ppm/deg]"
        sheet["H1"] = "Density [kg/m3]"
        sheet["I1"] = "Stress_x [MPa]"
        sheet["J1"] = "Poisson"
        sheet["K1"] = "R0"
        sheet["L1"] = "R"


        #Define a fill_color for cells
        fill_color = PatternFill(start_color="85c4f3", end_color="85c4f3", fill_type="solid")

        #Loop through the desired range of columns and rows to apply the fill_color and set a bold font
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=12):  # Adjust row/column range as needed
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = fill_color


        #Define a border style (thin lines for grid)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        #Apply the border to a range of cells and center the text in each cell
        for row in sheet.iter_rows(min_row=1, max_row=len(globals.materials)+1, min_col=1, max_col=12):  # Adjust range as needed
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        #Set the height and width values of cells in the excel file
        sheet.column_dimensions['A'].width = 13  #Width of column A
        sheet.column_dimensions['B'].width = 10  
        sheet.column_dimensions['C'].width = 6  
        sheet.column_dimensions['D'].width = 13  
        sheet.column_dimensions['E'].width = 10  
        sheet.column_dimensions['F'].width = 15  
        sheet.column_dimensions['G'].width = 15  
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15  
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10


        # Set row height
        # sheet.row_dimensions[1].height = 20  # Height of row 1 set to 30


        #Loop through materials{} and place values in excel file
        row_counter = 2

        for material in dict(reversed(globals.materials.items())):
            sheet.cell(row=row_counter, column=1, value=globals.materials[material]["Name"].get())
            sheet.cell(row=row_counter, column=2, value=globals.materials[material]["Thickness [nm]"].get())
            sheet.cell(row=row_counter, column=3, value=globals.materials[material]["Unit"].get())
            sheet.cell(row=row_counter, column=4, value=globals.materials[material]["Indent [nm]"].get())
            sheet.cell(row=row_counter, column=5, value=globals.materials[material]["Color"].get())
            sheet.cell(row=row_counter, column=6, value=globals.materials[material]["Modulus [GPa]"].get())
            sheet.cell(row=row_counter, column=7, value=globals.materials[material]["CTE [ppm/deg]"].get())
            sheet.cell(row=row_counter, column=8, value=globals.materials[material]["Density [kg/m3]"].get())
            sheet.cell(row=row_counter, column=9, value=globals.materials[material]["Stress_x [MPa]"].get())
            sheet.cell(row=row_counter, column=10, value=globals.materials[material]["Poisson"].get())
            sheet.cell(row=row_counter, column=11, value=globals.materials[material]["R0"].get())
            sheet.cell(row=row_counter, column=12, value=globals.materials[material]["R"].get())

            #increment row_counter
            row_counter += 1


        #Find coordinates of canvas on the screen
        canvas_x1 = globals.layer_stack_canvas.layer_stack_canvas.winfo_rootx()
        canvas_y1 = globals.layer_stack_canvas.layer_stack_canvas.winfo_rooty()
        canvas_x2 = canvas_x1 + globals.layer_stack_canvas.layer_stack_canvas.winfo_width()
        canvas_y2 = canvas_y1 + globals.layer_stack_canvas.layer_stack_canvas.winfo_height()

        bbox = (canvas_x1, canvas_y1, canvas_x2, canvas_y2)

        #Take a screenshot of the screen where canvas is
        screenshot = ImageGrab.grab(bbox=bbox)
        screenshot.save(f"{main_folder}/{sub_folder}/canvas_screenshot.png", "PNG")

        #Load the image with openpyxl
        canvas_screenshot = Image(f"{main_folder}/{sub_folder}/canvas_screenshot.png")

        #Set the width and height of image placed in excel file
        match globals.current_view:
            case "Stacked" | "Realistic" | "Stepped" | "Multi":
                canvas_screenshot.width = 350
                canvas_screenshot.height = 350
            

        #Add the image to the excel file in a specific cell
        sheet.add_image(canvas_screenshot, "N1")

        #Save the workbook as excel file
        workbook.save(f"{main_folder}/{sub_folder}/{filename}")
