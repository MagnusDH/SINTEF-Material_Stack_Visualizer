import tkinter
from tkinter import StringVar
import customtkinter
import settings
import globals
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from PIL import ImageGrab
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


#This class handles the buttons that perform actions on the canvas
class Canvas_Control_Panel:
    def __init__(self, window, row_placement:int, column_placement:int):
        # print("CANVAS_CONTROL_PANEL_INIT()")
        
        self.program_window = window

        #Row/column placement in main program window
        self.row_placement = row_placement
        self.column_placement = column_placement

        self.option_menu = None

        self.canvas_control_panel_frame = self.create_canvas_control_panel()


    def create_canvas_control_panel(self):
        """Creates a frame with widgets that performs actions on the layer_stack_canvas"""

        # print("CREATE_CANVAS_CONTROL_PANEL()")

        #if canvas_control_panel_frame has NOT been created before, create it
        if not hasattr(self, 'canvas_control_panel_frame'):
            #Create Frame from the control panel and place it within given window
            canvas_control_panel_frame = customtkinter.CTkFrame(
                master=self.program_window,
                fg_color=settings.canvas_control_panel_background_color
            )
            canvas_control_panel_frame.grid(
                row=self.row_placement,
                column=self.column_placement,
                padx=(settings.canvas_control_panel_padding_left, settings.canvas_control_panel_padding_right),
                pady=(settings.canvas_control_panel_padding_top, settings.canvas_control_panel_padding_bottom),
                sticky="nsew"
            )

            #Define the row&column layout of the material_control_panel_frame
            canvas_control_panel_frame.columnconfigure(0, weight=33, uniform="group1")
            canvas_control_panel_frame.columnconfigure(1, weight=33, uniform="group1")
            canvas_control_panel_frame.columnconfigure(2, weight=33, uniform="group1")

            canvas_control_panel_frame.rowconfigure(0, weight=50, uniform="group1")
            canvas_control_panel_frame.rowconfigure(1, weight=50, uniform="group1") 


        #Export Option menu label
        export_option_menu_label = customtkinter.CTkLabel(
            master=canvas_control_panel_frame,
            text="Export", 
            text_color=settings.canvas_control_panel_text_color,
            font=(settings.text_font, settings.canvas_control_panel_text_size, "bold"),
            bg_color=settings.canvas_control_panel_background_color
        )
        export_option_menu_label.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(5,5),
            pady=(5,5)
        )

        #Export option menu
        self.export_option_menu = customtkinter.CTkOptionMenu(
            master=canvas_control_panel_frame, 
            values=["Stack with text (SVG)", "Stack without text (SVG)", "Individual layers (SVG)", "Graphs (SVG)", "Material properties + screenshot (EXCEL)", "All"],
            font=(settings.text_font, settings.canvas_control_panel_text_size),
            bg_color=settings.canvas_control_panel_background_color,
            fg_color=settings.canvas_control_panel_dropdown_color,
            text_color=settings.canvas_control_panel_dropdown_text_color,
            button_color=settings.canvas_control_panel_dropdown_button_color,
            button_hover_color=settings.canvas_control_panel_dropdown_button_hover_color,
            dropdown_text_color=settings.canvas_control_panel_dropdown_text_color,
            dropdown_fg_color=settings.canvas_control_panel_dropdown_background_color,
            dropdown_hover_color=settings.canvas_control_panel_dropdown_hover_color,
            command=self.choose_stack_export
        )
        self.export_option_menu.grid(
            row=1, 
            column=0, 
            sticky="nsew",
            padx=(5,5), 
            pady=(5,5)
        )

        #Reset canvas button
        reset_canvas_button = customtkinter.CTkButton(
            master=canvas_control_panel_frame, 
            text="Reset canvas", 
            font=(settings.text_font, settings.canvas_control_panel_text_size),
            fg_color=settings.canvas_control_panel_button_color, 
            hover_color=settings.canvas_control_panel_button_hover_color, 
            text_color=settings.canvas_control_panel_button_text_color,
            command=self.reset_canvas
        )
        reset_canvas_button.grid(
            row=1, 
            column=1, 
            sticky="nsew", 
            padx=(5,5), 
            pady=(5,5)
        )    

        #Option menu "view" label
        option_menu_view_label = customtkinter.CTkLabel(
            master=canvas_control_panel_frame,
            text="View", 
            text_color=settings.canvas_control_panel_text_color,
            font=(settings.text_font, settings.canvas_control_panel_text_size, "bold"),
            # font=(settings.text_font, 15, "bold")
            bg_color=settings.canvas_control_panel_background_color
        )
        option_menu_view_label.grid(
            row=0,
            column=2,
            sticky="nsew",
            padx=(5,5),
            pady=(5,5)
        )

        #Switch layout option menu
        self.option_menu = customtkinter.CTkOptionMenu(
            master=canvas_control_panel_frame, 
            values=["Stacked", "Realistic", "Stepped", "Multi"],
            variable=globals.current_view,
            font=(settings.text_font, settings.canvas_control_panel_text_size),
            bg_color=settings.canvas_control_panel_background_color,
            fg_color=settings.canvas_control_panel_dropdown_color,
            text_color=settings.canvas_control_panel_dropdown_text_color,
            button_color=settings.canvas_control_panel_dropdown_button_color,
            button_hover_color=settings.canvas_control_panel_dropdown_button_hover_color,
            dropdown_text_color=settings.canvas_control_panel_dropdown_text_color,
            dropdown_fg_color=settings.canvas_control_panel_dropdown_background_color,
            dropdown_hover_color=settings.canvas_control_panel_dropdown_hover_color,
            # command=self.switch_layout
        )
        self.option_menu.grid(
            row=1, 
            column=2, 
            sticky="nsew",
            padx=(5,5), 
            pady=(5,5)
        )


        return canvas_control_panel_frame

    
    def reset_canvas(self):
        """Resets the scale of items and the start drawing point on the canvas back to their original scale and place"""
        
        # print("CLASS CANVAS_CONTROL_PANEL -> RESET_CANVAS()")

        #Find the original scale of the canvas
        inverse_scale = 1.0 / globals.layer_stack_canvas.current_scale

        #Scale all items back to their original size
        globals.layer_stack_canvas.layer_stack_canvas.scale("all", 0, 0, inverse_scale, inverse_scale)

        #Reset the original scale
        globals.layer_stack_canvas.current_scale = 1.0

        #Move the canvas back to its original drawing point
        globals.layer_stack_canvas.layer_stack_canvas.xview_moveto(0)
        globals.layer_stack_canvas.layer_stack_canvas.yview_moveto(0)

        #Redraw material stack
        globals.layer_stack_canvas.draw_material_stack()


    def choose_stack_export(self, *event):
        """Calls different export functions based on the 'export_option_menu' value"""
       
        # print("CHOOSE_STACK_EXPORT()")

        #Call functions based on export_option_menu value
        match self.export_option_menu.get():
            case "Stack with text (SVG)":
                #########
                self.program_window.geometry(f"{1500}x{300}")
                self.program_window.update()
                self.program_window.update_idletasks()
                #########
                self.export_full_stack_as_svg()

                #########
                self.program_window.geometry(f"{settings.program_window_width}x{settings.program_window_height}")
                self.program_window.update()
                self.program_window.update_idletasks()
                #########

            case "Stack without text (SVG)":
                self.export_material_stack_as_svg()

            case "Individual layers (SVG)":
                self.export_layers_as_svg()

            case "Graphs (SVG)":
                self.export_graphs()

            case "Material properties + screenshot (EXCEL)":
                #Make graphs if they have not been made
                if(globals.graph_canvas != None):
                    self.export_graphs()

                self.export_to_excel()

            case "All":
                self.export_full_stack_as_svg()
                self.export_material_stack_as_svg()
                self.export_layers_as_svg()
                self.export_graphs()
                self.export_to_excel()


    def export_material_stack_as_svg(self):
        """Exports only the material rectangles from the stack as SVG file"""

        # print("EXPORT_MATERIAL_STACK_AS_SVG")
        all_items = globals.layer_stack_canvas.layer_stack_canvas.find_all()

        #Return if there are no items on canvas
        if(len(all_items)-1 == 0):
            return

        #Create folders and filenames
        main_folder = "exports"
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder if it does not already exist
        sub_folder = globals.current_view.get()
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        
        #Create path for file to be saved in
        filename = f"{globals.current_view.get()}_stack_without_text.svg"
        file_path = os.path.join(f"{main_folder}/{sub_folder}/{filename}")

        #Open the file for writing
        with open(file_path, 'w') as f:
            #Write XML declaration for the SVG file, specifying the XML version, character encoding, and standalone status.
            f.write('<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n')

            #Write opening tag for the SVG file, specifying the width and height attributes based on the canvas dimensions. The xmlns attribute defines the XML namespace for SVG.
            f.write('<svg width="{}" height="{}" xmlns="http://www.w3.org/2000/svg">\n'.format(self.program_window.winfo_width(), self.program_window.winfo_height()))
 
            #Iterate through all the items
            for item in all_items:
                tag = globals.layer_stack_canvas.layer_stack_canvas.gettags(item)
                if(tag):
                    match tag[0]:
                        case "material_rectangle":
                            #Create SVG-element of material rectangle
                            rect_x0, rect_y0, rect_x1, rect_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            rectangle_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")
                            if(rectangle_color == ""):
                                rectangle_color = "white"
                            svg_rectangle_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, rectangle_color)
                            
                            f.write(svg_rectangle_element)

                            #Create SVG-element for the rectangle bounding box and write it to file
                            svg_bbox_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, settings.layer_stack_canvas_rectangle_outline_color)
                            f.write(svg_bbox_element)
            
            #Write the closing SVG tag to the file, completing the SVG file
            f.write('</svg>\n')

        #Close the svg file
        f.close()

   
    def export_layers_as_svg(self):
        """Exports every layer of the stack with text and arrows as SVG-file"""     
        # print("EXPORT_LAYERS_AS_SVG()")

        all_items = globals.layer_stack_canvas.layer_stack_canvas.find_all()

        #Return if there are no items on canvas
        if(len(all_items)-1 == 0):
            return

        #Create folders and filenames
        main_folder = "exports"
        #Create the folder if it doesn't exist
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder if it does not already exist
        sub_folder = globals.current_view.get()
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        

        #Each SVG-file is assigned a number based on how many layers are in each file
        layer_counter = 1
        previously_created_elements = []

        #Iterate through all the materials
        for material in globals.materials:
            #Only create svg element if there is a rectangle
            if("Rectangle_id" in globals.materials[material]):

                #Create a name for the SVG file for the current layer
                # filename = f"{layer_counter}materials_{material}_{globals.current_view.get()}.svg"
                filename = f"{layer_counter}materials_{list(globals.materials)[0]}-{material}.svg"
                

                #Create the file path by joining the folder path and the filename
                file_path = os.path.join(f"{main_folder}/{sub_folder}/{filename}")

                #Open file for writing
                with open(file_path, 'w') as f:
                    #Write XML declaration for the SVG file, specifying the XML version, character encoding, and standalone status.
                    f.write('<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n')
                    #Write opening tag for the SVG file, specifying the width and height attributes based on the canvas dimensions. The xmlns attribute defines the XML namespace for SVG.
                    f.write('<svg width="{}" height="{}" xmlns="http://www.w3.org/2000/svg">\n'.format(self.program_window.winfo_width(), self.program_window.winfo_height())) 


                    #Write the previous created elements to the current file
                    if(len(previously_created_elements) != 0):
                        for element in previously_created_elements:
                            f.write(element)

                    #Create SVG-element of material rectangle
                    rect_x0, rect_y0, rect_x1, rect_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Rectangle_id"])

                    svg_rectangle_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, globals.materials[material]["Color"].get())
                    f.write(svg_rectangle_element)
                    previously_created_elements.append(svg_rectangle_element)

                    #Create SVG-element for the rectangle bounding box and write it to file
                    svg_bbox_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, settings.layer_stack_canvas_rectangle_outline_color)
                    f.write(svg_bbox_element)
                    previously_created_elements.append(svg_bbox_element)

                    #Create SVG-element for material text
                    if("Text_id" in globals.materials[material]):

                        text_x0, text_y0 = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Text_id"])
                        text_content = globals.layer_stack_canvas.layer_stack_canvas.itemcget(globals.materials[material]["Text_id"], 'text')
                        svg_text_element = '<text x="{}" y="{}" fill="{}" font-size="{}" font-weight="bold" dominant-baseline="middle" text-anchor="middle">{}</text>\n'.format(text_x0, text_y0, settings.text_color, settings.svg_text_size, text_content)
                        f.write(svg_text_element)
                        previously_created_elements.append(svg_text_element)

                    #Create SVG-element for text bounding box
                    if("Text_bbox_id" in globals.materials[material]):

                        bbox_x0, bbox_y0, bbox_x1, bbox_y1 = globals.layer_stack_canvas.layer_stack_canvas.bbox(globals.materials[material]["Text_bbox_id"])
                        svg_bbox_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="black"/>\n'.format(bbox_x0, bbox_y0, bbox_x1-bbox_x0, bbox_y1-bbox_y0, settings.text_color)
                            
                        #Write the SVG representation of the bounding box to the file
                        f.write(svg_bbox_element)
                        previously_created_elements.append(svg_bbox_element)

                    #Create SVG-element for arrow line pointing from box to rectangle
                    if("Line_id" in globals.materials[material]):

                        #Line must be drawn from the right side of stack to left side of text
                        match globals.current_view.get():
                            case "Stacked" | "Realistic" | "Multi":
                                line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Line_id"])
                                #Construct an SVG <line> element for arrows
                                bbox_x0, bbox_y0, bbox_x1, bbox_y1 = globals.layer_stack_canvas.layer_stack_canvas.bbox(globals.materials[material]["Text_bbox_id"])
                                svg_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" marker-end="url(#arrow-end)" />\n'.format(bbox_x0, line_coords[1], line_coords[2]+7, line_coords[3], settings.text_color)

                                #Add arrowhead on the left side of the line
                                svg_line_element += (
                                '<defs>\n'
                                '    <marker id="arrow-end" markerWidth="10" markerHeight="10" refX="0" refY="3" orient="180">\n'
                                '        <path d="M0,0 L0,6 L9,3 z" fill="black" />\n'
                                '    </marker>\n'
                                '</defs>\n'
                                )

                                #Write the SVG representation of the arrow to the file
                                f.write(svg_line_element)
                                previously_created_elements.append(svg_line_element)


                            #Line must be drawn from the left side of stack to right side of text
                            case "Stepped":
                                line_x0, line_y0, line_x1, line_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Line_id"])
                                #Construct an SVG <line> element for arrows
                                bbox_x0, bbox_y0, bbox_x1, bbox_y1 = globals.layer_stack_canvas.layer_stack_canvas.bbox(globals.materials[material]["Text_bbox_id"])
                                svg_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" marker-end="url(#arrow-end)" />\n'.format(bbox_x1, line_y0, line_x1-7, line_y1, settings.text_color)
                                
                                #Add arrowhead on right side of the line
                                svg_line_element += (
                                    '<defs>\n'
                                    '    <marker id="arrow-end" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="180">\n'
                                    '        <path d="M0,3 L9,0 L9,6 z" fill="black" />\n'
                                    '    </marker>\n'
                                    '</defs>\n'
                                )

                                #Write the SVG representation of the arrow to the file
                                f.write(svg_line_element)
                                previously_created_elements.append(svg_line_element)


                    #Create SVG-element for indent_text
                    if("Indent_text_id" in globals.materials[material]):                    

                        indent_text_x0, indent_text_y0 = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Indent_text_id"])
                        indent_text_content = globals.layer_stack_canvas.layer_stack_canvas.itemcget(globals.materials[material]["Indent_text_id"], 'text')
                        svg_indent_text_element = '<text x="{}" y="{}" fill="black" font-size="{}" font-weight="bold" dominant-baseline="middle" text-anchor="middle">{}</text>\n'.format(indent_text_x0, indent_text_y0, settings.svg_text_size, indent_text_content)
                            
                        f.write(svg_indent_text_element)
                        previously_created_elements.append(svg_indent_text_element)


                    #Create SVG-element for indent_text bounding box
                    if("Indent_text_bbox_id" in globals.materials[material]):                    

                        bbox_x0, bbox_y0, bbox_x1, bbox_y1 = globals.layer_stack_canvas.layer_stack_canvas.bbox(globals.materials[material]["Indent_text_bbox_id"])
                        svg_indent_text_bbox_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="black"/>\n'.format(bbox_x0, bbox_y0, bbox_x1-bbox_x0, bbox_y1-bbox_y0, settings.text_color)
                            
                        #Write the SVG representation of the bounding box to the file
                        f.write(svg_indent_text_bbox_element)
                        previously_created_elements.append(svg_indent_text_bbox_element)


                    #Create SVG-element for indent_line
                    if("Indent_line_id" in globals.materials[material]):                    

                        indent_line_x0, indent_line_y0, indent_line_x1, indent_line_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Indent_line_id"])
                        #Construct an SVG <line> element for arrows
                        svg_indent_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="black" marker-start="url(#arrow-start)" marker-end="url(#arrow-end)" />\n'.format(indent_line_x0+8, indent_line_y0, indent_line_x1-10, indent_line_y1)
                        
                        #Add arrowheads on both sides of the line
                        svg_indent_line_element += (
                            '<defs>\n'
                            '    <marker id="arrow-end" markerWidth="10" markerHeight="10" refX="0" refY="3" orient="0">\n'
                            '        <path d="M0,0 L0,6 L9,3 z" fill="black" />\n'
                            '    </marker>\n'
                            '    <marker id="arrow-start" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="0">\n'
                            '        <path d="M0,3 L9,0 L9,6 z" fill="black" />\n'
                            '    </marker>\n'
                            '</defs>\n'
                        )
                        
                        #Write the SVG representation of the arrow to the file
                        f.write(svg_indent_line_element)
                        previously_created_elements.append(svg_indent_line_element)


                    #Create SVG-element for arrow line pointing from indent_text to indent_line
                    if("Indent_arrow_pointer_id" in globals.materials[material]):                    

                        line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(globals.materials[material]["Indent_arrow_pointer_id"])
                        #Construct an SVG <line> element for arrows
                        bbox_x0, bbox_y0, bbox_x1, bbox_y1 = globals.layer_stack_canvas.layer_stack_canvas.bbox(globals.materials[material]["Indent_text_bbox_id"])
                        svg_indent_arrow_pointer_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" marker-end="url(#arrow-start)" />\n'.format(bbox_x0, line_coords[1], line_coords[2]+7, line_coords[3], settings.text_color)

                        #Add arrowhead on the left side of the line
                        svg_indent_arrow_pointer_element += (
                            '<defs>\n'
                            '    <marker id="arrow-end" markerWidth="10" markerHeight="10" refX="0" refY="3" orient="180">\n'
                            '        <path d="M0,0 L0,6 L9,3 z" fill="black" />\n'
                            '    </marker>\n'
                            '</defs>\n'
                        )

                        #Write the SVG representation of the arrow to the file
                        f.write(svg_indent_arrow_pointer_element)
                        previously_created_elements.append(svg_indent_arrow_pointer_element)


                    #Write the closing SVG tag to the file, completing the SVG file
                    f.write('</svg>\n')

                #Close the svg file
                f.close()

                #Increment layer_counter
                layer_counter += 1
        

    def export_full_stack_as_svg(self):
        """Exports all details on the stack as SVG-file"""
        # print("EXPORT_FULL_STACK_AS_SVG")
        all_items = globals.layer_stack_canvas.layer_stack_canvas.find_all()

        #Return if there are no items on canvas
        if(len(all_items)-1 == 0):
            return

        #Create folders and filenames
        main_folder = "exports"
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder if it does not already exist
        sub_folder = globals.current_view.get()
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        
        #Create path for file to be saved in
        filename = f"{globals.current_view.get()}_stack_with_text.svg"
        file_path = os.path.join(f"{main_folder}/{sub_folder}/{filename}")


        #Open the file for writing
        with open(file_path, 'w') as f:
            #Write XML declaration for the SVG file, specifying the XML version, character encoding, and standalone status.
            f.write('<?xml version="1.0" encoding="UTF-8" standalone="no"?>\n')

            #Write opening tag for the SVG file, specifying the width and height attributes based on the canvas dimensions. The xmlns attribute defines the XML namespace for SVG.
            f.write(f"<svg width='{globals.layer_stack_canvas.layer_stack_canvas.winfo_width()}' height='{globals.layer_stack_canvas.layer_stack_canvas.winfo_height()}' xmlns='http://www.w3.org/2000/svg'>\n")
           

            #Iterate through all the items
            for item in all_items:
                tag = globals.layer_stack_canvas.layer_stack_canvas.gettags(item)
                if(tag):
                    match tag[0]:
                        case "layer_stack_canvas_bounding_rectangle":
                            continue

                        case "material_rectangle":
                            #Create SVG-element of material rectangle
                            rect_x0, rect_y0, rect_x1, rect_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            rectangle_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")
                            if(rectangle_color == ""):
                                rectangle_color = "white"
                            svg_rectangle_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, rectangle_color)
                            
                            f.write(svg_rectangle_element)

                            #Create SVG-element for the rectangle bounding box and write it to file
                            svg_bbox_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="{}" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0, settings.layer_stack_canvas_rectangle_outline_color)
                            f.write(svg_bbox_element)

                        case "text":
                            #Create SVG-element for text
                            text_x0, text_y0 = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            text_content = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, 'text')
                            text_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, 'fill')
                            svg_text_element = '<text x="{}" y="{}" fill="{}" font-size="{}" font-weight="bold" dominant-baseline="middle" text-anchor="middle">{}</text>\n'.format(text_x0, text_y0, text_color, settings.svg_text_size, text_content)
                            f.write(svg_text_element)
                            
                        case "text_bbox":
                            #Create SVG-element of text bbox
                            rect_x0, rect_y0, rect_x1, rect_y1 = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            svg_rectangle_element = '<rect x="{}" y="{}" width="{}" height="{}" fill="none" stroke="black" stroke-width="1" />\n'.format(rect_x0, rect_y0, rect_x1 - rect_x0, rect_y1 - rect_y0)
                            
                            f.write(svg_rectangle_element)
                        
                        case "line":
                            #Create line SVG-element
                            line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            line_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")
                            svg_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="1" />\n'.format(line_coords[0], line_coords[1], line_coords[2], line_coords[3], line_color)

                            f.write(svg_line_element)

                        case "dotted_line":
                            #Create line SVG-element
                            line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            line_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")
                            svg_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="2" stroke-dasharray="5,5" />\n'.format(line_coords[0], line_coords[1], line_coords[2], line_coords[3], line_color)

                            f.write(svg_line_element)

                        case "arrow_line":
                            line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            line_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")

                            svg_line_element = '<line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" marker-end="url(#arrow-end)" />\n'.format(line_coords[0], line_coords[1], line_coords[2], line_coords[3], line_color)

                            #Add arrowhead on top of line
                            svg_line_element += (
                            '<defs>\n'
                            '    <marker id="arrow-end" markerWidth="10" markerHeight="10" refX="10" refY="3" orient="auto">\n'
                            '        <path d="M0,0 L0,6 L9,3 z" fill="black" />\n'
                            '    </marker>\n'
                            '</defs>\n'
                            )

                            f.write(svg_line_element)

                        case "arrow_line_both":
                            line_coords = globals.layer_stack_canvas.layer_stack_canvas.coords(item)
                            line_color = globals.layer_stack_canvas.layer_stack_canvas.itemcget(item, "fill")

                            #Use a unique ID based on the line color
                            color_id = line_color.replace("#", "")  # Strip # for ID safety

                            #SVG line element
                            svg_line_element = (
                                f'<line x1="{line_coords[0]}" y1="{line_coords[1]}" '
                                f'x2="{line_coords[2]}" y2="{line_coords[3]}" '
                                f'stroke="{line_color}" stroke-width="2" '
                                f'marker-start="url(#arrow-start-{color_id})" '
                                f'marker-end="url(#arrow-end-{color_id})" />\n'
                            )

                            #Add arrowheads on both sides of the line with the same color
                            svg_line_element += (
                                f'<defs>\n'
                                f'    <marker id="arrow-end-{color_id}" markerWidth="10" markerHeight="10" refX="10" refY="3" orient="auto">\n'
                                f'        <path d="M0,0 L0,6 L9,3 z" fill="{line_color}" />\n'
                                f'    </marker>\n'
                                f'    <marker id="arrow-start-{color_id}" markerWidth="10" markerHeight="10" refX="0" refY="3" orient="auto">\n'
                                f'        <path d="M0,3 L9,0 L9,6 z" fill="{line_color}" />\n'
                                f'    </marker>\n'
                                f'</defs>\n'
                            )

                            f.write(svg_line_element)


            #Write the closing SVG tag to the file, completing the SVG file
            f.write('</svg>\n')

        #Close the svg file
        f.close()


    def export_graphs(self):
        """Exports the graph as svg file"""
        
        # print("EXPORT_GRAPHS()")

        #CREATE FOLDER HIERARCHY
        main_folder = "exports"

        #Create folder if it does not exist
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder
        sub_folder = "graphs"

        #Create sub_folder if it does not exist
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        
        #Create name for the file
        filenameSVG = "graphs.svg"
        filenameJPG = "graphs.jpg"


        #Save the graph as .svg and .jpg file
        globals.graph_canvas.graph_container.savefig(f"{main_folder}/{sub_folder}/{filenameSVG}")
        globals.graph_canvas.graph_container.savefig(f"{main_folder}/{sub_folder}/{filenameJPG}")


    def export_to_excel(self):
        """
        Creates two tabs in an excel file containing information about each material, equations, pictures of the stack and the graphs
        """

        # print("EXPORT_TO_EXCEL()")

        #Create an filename and a workbook to contain data
        filename = "exported_materials.xlsx"

        #Create main folder if it does not already exist
        main_folder = "exports"
        if not os.path.exists(main_folder):
            os.makedirs(main_folder)

        #Create sub_folder if it does not already exist
        sub_folder = "excel"
        if not os.path.exists(f"{main_folder}/{sub_folder}"):
            os.makedirs(f"{main_folder}/{sub_folder}")
        
        #Update all equations so that correct numbers are put in excel file
        globals.app.calculate_all_equations()

        #Create excel workbook/file
        workbook = Workbook()

        #Create "calculations" tab in excel file
        self.create_excel_materials_tab(workbook)

        #Create "calculations" tab in excel file
        self.create_excel_calculations_tab(workbook)

        #Save the workbook as excel file
        workbook.save(f"{main_folder}/{sub_folder}/{filename}")

        
    def create_excel_materials_tab(self, excel_workbook):
        # print("CREATE_EXCEL_MATERIALS_TAB()")

        main_folder = "exports"
        sub_folder = "excel"

        tab = excel_workbook.active
        tab.title = "Materials"

        #Create header cells
        tab["A1"] = "Material"          # Add a new header in column A row 1
        tab["B1"] = "Thickness [nm]"
        tab["C1"] = "Unit"
        tab["D1"] = "Indent [nm]"
        tab["E1"] = "Color"
        tab["F1"] = "Modulus [GPa]"
        tab["G1"] = "CTE [ppm/deg]"
        tab["H1"] = "Density [kg/m3]"
        tab["I1"] = "Stress_x [MPa]"
        tab["J1"] = "Poisson"
        tab["K1"] = "R0"
        tab["L1"] = "R"


        #Define a fill_color for cells
        fill_color = PatternFill(start_color="2f9ff5", end_color="2f9ff5", fill_type="solid")

        #Loop through the desired range of columns and rows to apply the fill_color and set a bold font
        for row in tab.iter_rows(min_row=1, max_row=1, min_col=1, max_col=12):  # Adjust row/column range as needed
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = fill_color


        #Define a border style (thin lines for grid)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        #Apply the border to a range of cells and center the text in each cell
        for row in tab.iter_rows(min_row=1, max_row=len(globals.materials)+1, min_col=1, max_col=12):  # Adjust range as needed
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        #Set the height and width values of cells in the excel file
        tab.column_dimensions['A'].width = 13  #Width of column A
        tab.column_dimensions['B'].width = 10  
        tab.column_dimensions['C'].width = 6  
        tab.column_dimensions['D'].width = 13  
        tab.column_dimensions['E'].width = 10  
        tab.column_dimensions['F'].width = 15  
        tab.column_dimensions['G'].width = 15  
        tab.column_dimensions['H'].width = 15
        tab.column_dimensions['I'].width = 15  
        tab.column_dimensions['J'].width = 10
        tab.column_dimensions['K'].width = 10
        tab.column_dimensions['L'].width = 10


        #Loop through materials{} and place values in excel file
        row_counter = 2

        for material in dict(reversed(globals.materials.items())):
            tab.cell(row=row_counter, column=1, value=globals.materials[material]["Name"].get())
            tab.cell(row=row_counter, column=2, value=globals.materials[material]["Thickness [nm]"].get())
            tab.cell(row=row_counter, column=3, value=globals.materials[material]["Unit"].get())
            tab.cell(row=row_counter, column=4, value=globals.materials[material]["Indent [nm]"].get())
            tab.cell(row=row_counter, column=5, value=globals.materials[material]["Color"].get())
            tab.cell(row=row_counter, column=6, value=globals.materials[material]["Modulus [GPa]"].get())
            tab.cell(row=row_counter, column=7, value=globals.materials[material]["CTE [ppm/deg]"].get())
            tab.cell(row=row_counter, column=8, value=globals.materials[material]["Density [kg/m3]"].get())
            tab.cell(row=row_counter, column=9, value=globals.materials[material]["Stress_x [MPa]"].get())
            tab.cell(row=row_counter, column=10, value=globals.materials[material]["Poisson"].get())
            tab.cell(row=row_counter, column=11, value=globals.materials[material]["R0"].get())
            tab.cell(row=row_counter, column=12, value=globals.materials[material]["R"].get())

            #increment row_counter
            row_counter += 1


        #Find coordinates of canvas on the screen
        canvas_x1 = globals.layer_stack_canvas.layer_stack_canvas.winfo_rootx()
        canvas_y1 = globals.layer_stack_canvas.layer_stack_canvas.winfo_rooty()
        canvas_x2 = canvas_x1 + globals.layer_stack_canvas.layer_stack_canvas.winfo_width()
        canvas_y2 = canvas_y1 + globals.layer_stack_canvas.layer_stack_canvas.winfo_height()

        bbox = (canvas_x1, canvas_y1, canvas_x2, canvas_y2)


        #########test########
        globals.layer_stack_canvas.layer_stack_canvas.delete("all")
        for material in globals.materials:
            if("Rectangle_id" in globals.materials[material]):
                del globals.materials[material]["Rectangle_id"]
            
            if("Text_id" in globals.materials[material]):
                del globals.materials[material]["Text_id"]
    
            if("Text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Text_bbox_id"]
            
            if("Line_id" in globals.materials[material]):
                del globals.materials[material]["Line_id"]
            
            if("Indent_text_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_id"]
            
            if("Indent_text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_bbox_id"]
            
            if("Indent_line_id" in globals.materials[material]):
                del globals.materials[material]["Indent_line_id"]
            
            if("Indent_arrow_pointer_id" in globals.materials[material]):
                del globals.materials[material]["Indent_arrow_pointer_id"]

        # Sort the materials dictionary after the "layer" value
        globals.materials = dict(sorted(globals.materials.items(), key=lambda item: item[1]["Layer"].get()))
        globals.layer_stack_canvas.draw_material_stack_stacked()        
        globals.layer_stack_canvas.write_text_on_stack()
        self.program_window.update()
        screenshot = ImageGrab.grab(bbox=bbox)
        screenshot.save(f"{main_folder}/{sub_folder}/canvas_screenshot_stacked.png", "PNG")


        #Take a screenshot of the screen where canvas is
        globals.layer_stack_canvas.layer_stack_canvas.delete("all")
        for material in globals.materials:
            if("Rectangle_id" in globals.materials[material]):
                del globals.materials[material]["Rectangle_id"]
            
            if("Text_id" in globals.materials[material]):
                del globals.materials[material]["Text_id"]
    
            if("Text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Text_bbox_id"]
            
            if("Line_id" in globals.materials[material]):
                del globals.materials[material]["Line_id"]
            
            if("Indent_text_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_id"]
            
            if("Indent_text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_bbox_id"]
            
            if("Indent_line_id" in globals.materials[material]):
                del globals.materials[material]["Indent_line_id"]
            
            if("Indent_arrow_pointer_id" in globals.materials[material]):
                del globals.materials[material]["Indent_arrow_pointer_id"]

        #Sort the materials dictionary after the "layer" value
        globals.materials = dict(sorted(globals.materials.items(), key=lambda item: item[1]["Layer"].get()))
        globals.layer_stack_canvas.draw_material_stack_realistic()        
        self.program_window.update()
        screenshot = ImageGrab.grab(bbox=bbox)
        screenshot.save(f"{main_folder}/{sub_folder}/canvas_screenshot_realistic.png", "PNG")



        #Take a screenshot of the screen where canvas is
        globals.layer_stack_canvas.layer_stack_canvas.delete("all")
        for material in globals.materials:
            if("Rectangle_id" in globals.materials[material]):
                del globals.materials[material]["Rectangle_id"]
            
            if("Text_id" in globals.materials[material]):
                del globals.materials[material]["Text_id"]
    
            if("Text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Text_bbox_id"]
            
            if("Line_id" in globals.materials[material]):
                del globals.materials[material]["Line_id"]
            
            if("Indent_text_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_id"]
            
            if("Indent_text_bbox_id" in globals.materials[material]):
                del globals.materials[material]["Indent_text_bbox_id"]
            
            if("Indent_line_id" in globals.materials[material]):
                del globals.materials[material]["Indent_line_id"]
            
            if("Indent_arrow_pointer_id" in globals.materials[material]):
                del globals.materials[material]["Indent_arrow_pointer_id"]

        #Sort the materials dictionary after the "layer" value
        globals.materials = dict(sorted(globals.materials.items(), key=lambda item: item[1]["Layer"].get()))
        globals.layer_stack_canvas.draw_material_stack_realistic()
        globals.layer_stack_canvas.write_text_on_stack()
        globals.layer_stack_canvas.draw_Zn_and_Zp()   
        self.program_window.update()
        screenshot = ImageGrab.grab(bbox=bbox)
        screenshot.save(f"{main_folder}/{sub_folder}/canvas_screenshot_multi.png", "PNG")
        #####################

        #Load the image with openpyxl
        canvas_screenshot_stacked = Image(f"{main_folder}/{sub_folder}/canvas_screenshot_stacked.png")
        canvas_screenshot_realistic = Image(f"{main_folder}/{sub_folder}/canvas_screenshot_realistic.png")
        canvas_screenshot_multi = Image(f"{main_folder}/{sub_folder}/canvas_screenshot_multi.png")


        #Set the width and height of image placed in excel file
        canvas_screenshot_stacked.width = 350
        canvas_screenshot_stacked.height = 450

        canvas_screenshot_realistic.width = 350
        canvas_screenshot_realistic.height = 450

        canvas_screenshot_multi.width = 350
        canvas_screenshot_multi.height = 450
            

        #Add the image to the excel file in a specific cell
        tab.add_image(canvas_screenshot_stacked, "N1")
        tab.add_image(canvas_screenshot_realistic, "T1")
        tab.add_image(canvas_screenshot_multi, "Z1")


    def create_excel_calculations_tab(self, excel_workbook):
        """
        Creates a 'calculations' tab in the given excel file
        """
      
        # print("CREATE_EXCEL_CALCULATIONS_TAB()")

        tab = excel_workbook.create_sheet(title="Calculations")

        #Set the width value of cells in the excel file
        tab.column_dimensions['A'].width = 25
        tab.column_dimensions['B'].width = 15


        #Create cells
        tab.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        tab["A1"] = "Parameters"
        tab["A1"].font = Font(bold=True)
        tab["A1"].alignment = Alignment(horizontal="center", vertical="center")
        tab["A1"].fill = PatternFill(start_color="2f9ff5", end_color="2f9ff5", fill_type="solid")


        tab["A2"] = "e₃₁ [C/m²]"
        tab["B2"] = globals.e_31_f_value.get()
        
        tab["A3"] = "Zn [m]"
        #Insert Zn as "meters"
        tab["B3"] = globals.Zn.get()

        tab["A4"] = "Curve [1/m]"
        tab["B4"] = globals.curv_is.get()

        tab["A5"] = "EI_is [Nm²]"
        tab["B5"] = globals.EI.get()

        tab["A6"] = "M_is [nm]"
        tab["B6"] = globals.M_is.get()

        tab["A7"] = "Stress neutral [nm]"
        tab["B7"] = "???"

        tab["A8"] = "L [μm]"
        tab["B8"] = globals.L_value.get()


        tab["A9"] = "Volt"
        tab["B9"] = globals.volt_value.get()


        #Apply borders around the cells
        start_row = 1
        end_row = 10
        start_column = 1
        end_column = 3
        thin_side = Side(style="thin")
        for row in range(start_row, end_row):
            for col in range(start_column, end_column):
                cell = tab.cell(row=row, column=col)
                cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


        tab.merge_cells(start_row=11, start_column=1, end_row=11, end_column=2)
        tab["A11"] = "Plot"
        tab["A11"].font = Font(bold=True)
        tab["A11"].alignment = Alignment(horizontal="center", vertical="center")
        tab["A11"].fill = PatternFill(start_color="2f9ff5", end_color="2f9ff5", fill_type="solid")



        #Create cells for each piezo material selected
        column_counter = 1
        row_counter = 12
        fill_color = PatternFill(start_color="93cefa", end_color="93cefa", fill_type="solid")

        for material in dict(reversed(globals.materials.items())):
            if(globals.materials[material]["Piezo_checkbox_id"].get() == "on"):
                
                #Set column dimensions
                if(column_counter % 2 != 0):
                    column_letter = get_column_letter(column_counter)
                    tab.column_dimensions[column_letter].width = 25
                    


                #Merge cells for material name
                tab.merge_cells(start_row=row_counter, start_column=column_counter, end_row=row_counter, end_column=column_counter+1)
                header_cell = tab.cell(row=row_counter, column=column_counter, value=f"{material} (layer {globals.materials[material]["Layer"].get()})")
                header_cell.font = Font(bold=True)
                header_cell.fill = fill_color

                tab.cell(row=row_counter+1, column=column_counter, value="ZP [nm]")
                #Insert Zp in unit "nanometers"
                tab.cell(row=row_counter+1, column=column_counter+1, value=round(globals.materials[material]["Zp_value"].get()*1e9, 1))

                tab.cell(row=row_counter+2, column=column_counter, value="Mp [Nm]")
                tab.cell(row=row_counter+2, column=column_counter+1, value=globals.materials[material]["Mp_value"].get())
                tab.cell(row=row_counter+2, column=column_counter+1, value=f"{globals.materials[material]["Mp_value"].get():.2e}")


                tab.cell(row=row_counter+3, column=column_counter, value="Blocking force cantilever tip")
                tab.cell(row=row_counter+3, column=column_counter+1, value=f"{globals.materials[material]["Blocking_force_value"].get():.2e}")

                #Increment column counter
                row_counter += 4
            
        
        #Apply borders around the piezo material cells
        start_row = 11
        end_row = row_counter
        start_column = 1
        end_column = 3
        thin_side = Side(style="thin")
        for row in range(start_row, end_row):
            for col in range(start_column, end_column):
                cell = tab.cell(row=row, column=col)
                cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        #Align cells in column B to "left"
        left_alignment = Alignment(horizontal="left")
        for row in range(1, tab.max_row + 1):
            cell = tab.cell(row=row, column=2)  # column 2 is 'B'
            cell.alignment = left_alignment
        

        #Load graphs image
        main_folder = "exports"
        sub_folder = "graphs"
        filenameJPG = "graphs.jpg"
        graphs_image = Image(f"{main_folder}/{sub_folder}/{filenameJPG}")

        #Set the width and height of image placed in excel file
        graphs_image.width = 400
        graphs_image.height = 600

        #Add the image to the excel file in a specific cell
        tab.add_image(graphs_image, "E1")
